"""Engine de export XLSX (Sprint 4).

Gera planilha orçamentária ANF programaticamente (sem template externo) usando
openpyxl. Cinco abas:
  1. Resumo Executivo — projeto, área, custo direto, BDI, total final, R$/m²
  2. Planilha Orçamentária — EAP/WBS com fórmulas Excel ativas
  3. Composição BDI — quebra Acórdão TCU 2622/2013
  4. Levantamento Quantitativo — itens de ob_quantitativos
  5. Curva ABC — itens ordenados por peso decrescente, classes A/B/C destacadas
"""
from __future__ import annotations

import io
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..lib.supabase import get_supabase
from .bdi import calcular_bdi

logger = logging.getLogger(__name__)

# Estilos
_HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")  # brand-900
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TOTAL_FILL = PatternFill("solid", fgColor="F3F4F6")
_TOTAL_FONT = Font(bold=True, size=11)
_CURVA_A_FILL = PatternFill("solid", fgColor="FEE2E2")  # red-100
_CURVA_B_FILL = PatternFill("solid", fgColor="FEF3C7")  # amber-100
_CURVA_C_FILL = PatternFill("solid", fgColor="DCFCE7")  # green-100
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def render_orcamento_xlsx(project_id: str, bdi: dict[str, Any] | None = None) -> bytes:
    """Renderiza o XLSX completo. BDI default = Acórdão TCU típico."""
    sb = get_supabase()
    bdi = bdi or calcular_bdi()

    proj_res = sb.table("ob_projects").select("*").eq("id", project_id).limit(1).execute()
    if not proj_res.data:
        raise ValueError(f"projeto {project_id} não encontrado")
    project = proj_res.data[0]

    items_res = (
        sb.table("ob_orcamento_items")
        .select("*")
        .eq("project_id", project_id)
        .order("eap_code")
        .execute()
    )
    items = items_res.data or []

    quant_res = (
        sb.table("ob_quantitativos")
        .select("*")
        .eq("project_id", project_id)
        .order("disciplina")
        .execute()
    )
    quantitativos = quant_res.data or []

    wb = Workbook()
    # remove default sheet
    default_ws = wb.active
    if default_ws:
        wb.remove(default_ws)

    _aba_resumo(wb, project, items, bdi)
    _aba_planilha(wb, project, items, bdi)
    _aba_bdi(wb, bdi)
    _aba_levantamento(wb, quantitativos)
    _aba_curva_abc(wb, items)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Abas
# ─────────────────────────────────────────────────────────────────────────────

def _aba_resumo(
    wb: Workbook,
    project: dict[str, Any],
    items: list[dict[str, Any]],
    bdi: dict[str, Any],
) -> None:
    ws = wb.create_sheet("Resumo Executivo")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    rows: list[tuple[str, Any]] = [
        ("Projeto", project.get("name", "")),
        ("Tipo de obra", project.get("tipo_obra") or "—"),
        ("Localização", f"{project.get('cidade') or ''}/{project.get('uf') or ''}".strip("/") or "—"),
        ("Área total (m²)", project.get("area_total_m2") or 0),
        ("Status", project.get("status") or "—"),
        ("", ""),
        ("Total de itens", len(items)),
    ]

    custo_direto = sum(_to_float(it.get("custo_total")) for it in items)
    custo_total = custo_direto * float(bdi["multiplicador"])
    area = _to_float(project.get("area_total_m2"))

    rows.extend(
        [
            ("Custo direto (R$)", round(custo_direto, 2)),
            ("BDI (%)", bdi["bdi_pct"]),
            ("Multiplicador BDI", bdi["multiplicador"]),
            ("Custo total c/ BDI (R$)", round(custo_total, 2)),
            ("R$/m²", round(custo_total / area, 2) if area > 0 else "—"),
        ]
    )

    ws.cell(1, 1, "Resumo Executivo").font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells("A1:B1")

    for i, (label, value) in enumerate(rows, start=3):
        cell_label = ws.cell(i, 1, label)
        cell_value = ws.cell(i, 2, value)
        cell_label.font = Font(bold=True)
        cell_label.alignment = _LEFT
        cell_value.alignment = _LEFT
        if isinstance(value, (int, float)) and "R$" in label:
            cell_value.number_format = "#,##0.00"
        elif isinstance(value, (int, float)) and "%" in label:
            cell_value.number_format = "#,##0.0000"


def _aba_planilha(
    wb: Workbook,
    project: dict[str, Any],
    items: list[dict[str, Any]],
    bdi: dict[str, Any],
) -> None:
    ws = wb.create_sheet("Planilha Orçamentária")
    headers = [
        "EAP",
        "Descrição",
        "Unid",
        "Qtd",
        "Custo Unit (R$)",
        "Custo Total (R$)",
        "BDI %",
        "Custo Total c/ BDI (R$)",
        "Fonte",
        "Código",
        "Data base",
    ]
    widths = [12, 60, 8, 10, 14, 16, 8, 18, 12, 14, 12]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(1, col, h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    bdi_pct = float(bdi["bdi_pct"])
    bdi_mult = float(bdi["multiplicador"])

    for i, item in enumerate(items, start=2):
        ws.cell(i, 1, item.get("eap_code") or "")
        ws.cell(i, 2, item.get("descricao") or "")
        ws.cell(i, 3, item.get("unidade") or "")
        ws.cell(i, 4, _to_float(item.get("quantidade")))
        ws.cell(i, 5, _to_float(item.get("custo_unitario")))
        # Total = D*E (fórmula Excel)
        ws.cell(i, 6, f"=D{i}*E{i}")
        ws.cell(i, 7, bdi_pct / 100)
        # Total c/ BDI = F * (1 + G)  → mas mantemos consistente com multiplicador único
        ws.cell(i, 8, f"=F{i}*{bdi_mult}")
        ws.cell(i, 9, item.get("fonte") or "")
        ws.cell(i, 10, item.get("fonte_codigo") or "")
        ws.cell(i, 11, item.get("fonte_data_base") or "")

        # Highlight curva ABC classe A
        if item.get("curva_abc_classe") == "A":
            for col in range(1, len(headers) + 1):
                ws.cell(i, col).fill = _CURVA_A_FILL
        elif item.get("curva_abc_classe") == "B":
            for col in range(1, len(headers) + 1):
                ws.cell(i, col).fill = _CURVA_B_FILL

        for col in range(1, len(headers) + 1):
            ws.cell(i, col).border = _BORDER
        for col in (4, 5, 6, 7, 8):
            ws.cell(i, col).number_format = "#,##0.00" if col != 7 else "0.00%"
            ws.cell(i, col).alignment = _RIGHT

    # Totais ao fim
    total_row = len(items) + 3
    ws.cell(total_row, 1, "TOTAL").font = _TOTAL_FONT
    ws.cell(total_row, 1).fill = _TOTAL_FILL
    ws.cell(total_row, 6, f"=SUM(F2:F{total_row - 1})")
    ws.cell(total_row, 6).number_format = "#,##0.00"
    ws.cell(total_row, 6).font = _TOTAL_FONT
    ws.cell(total_row, 6).fill = _TOTAL_FILL
    ws.cell(total_row, 8, f"=SUM(H2:H{total_row - 1})")
    ws.cell(total_row, 8).number_format = "#,##0.00"
    ws.cell(total_row, 8).font = _TOTAL_FONT
    ws.cell(total_row, 8).fill = _TOTAL_FILL

    ws.freeze_panes = "B2"
    # Avisa o user
    info_row = total_row + 2
    ws.cell(
        info_row,
        1,
        f"Projeto: {project.get('name', '')} · BDI {bdi_pct:.2f}% (multiplicador {bdi_mult:.4f})",
    )
    ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=8)


def _aba_bdi(wb: Workbook, bdi: dict[str, Any]) -> None:
    ws = wb.create_sheet("Composição BDI")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    ws.cell(1, 1, "Composição do BDI (Acórdão TCU 2622/2013)").font = Font(
        bold=True, size=14, color="1E3A8A"
    )
    ws.merge_cells("A1:B1")

    rows: list[tuple[str, float | str]] = [
        ("Lucro (%)", bdi["lucro_pct"]),
        ("Despesas indiretas / Adm central (%)", bdi["despesas_indiretas_pct"]),
        ("Risco (%)", bdi["risco_pct"]),
        ("Despesas financeiras (%)", bdi["despesas_financeiras_pct"]),
        ("", ""),
        ("ISS (%)", bdi["tributos"]["iss"]),
        ("PIS (%)", bdi["tributos"]["pis"]),
        ("COFINS (%)", bdi["tributos"]["cofins"]),
        ("IRPJ (%)", bdi["tributos"]["irpj"]),
        ("CSLL (%)", bdi["tributos"]["csll"]),
        ("Total tributos (%)", bdi["tributos"]["total"]),
        ("", ""),
        ("AC (fração)", bdi["componentes_fracao"]["AC"]),
        ("DF (fração)", bdi["componentes_fracao"]["DF"]),
        ("L (fração)", bdi["componentes_fracao"]["L"]),
        ("I (fração)", bdi["componentes_fracao"]["I"]),
        ("", ""),
        ("BDI (%)", bdi["bdi_pct"]),
        ("Multiplicador (1 + BDI)", bdi["multiplicador"]),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(i, 1, label).font = Font(bold=label.startswith("BDI") or label.startswith("Multipl"))
        ws.cell(i, 1).alignment = _LEFT
        c = ws.cell(i, 2, value)
        c.alignment = _RIGHT
        if isinstance(value, (int, float)):
            if "fração" in label.lower() or "Multipl" in label:
                c.number_format = "#,##0.000000"
            elif "%" in label:
                c.number_format = "#,##0.0000"
        if "BDI (%)" in label:
            c.font = Font(bold=True, size=12, color="1E3A8A")

    ws.cell(len(rows) + 5, 1, "Fórmula: BDI = ((1+AC)·(1+DF)·(1+L)) / (1-I) - 1").font = Font(italic=True)
    ws.merge_cells(start_row=len(rows) + 5, start_column=1, end_row=len(rows) + 5, end_column=2)


def _aba_levantamento(wb: Workbook, quantitativos: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Levantamento Quantitativo")
    headers = [
        "Disciplina",
        "Código",
        "Descrição",
        "Unid",
        "Qtd",
        "Memorial de cálculo",
        "Origem",
        "Confiança",
        "Revisar?",
    ]
    widths = [12, 14, 45, 8, 10, 50, 25, 12, 10]
    _write_table_header(ws, headers, widths)

    for i, q in enumerate(quantitativos, start=2):
        ws.cell(i, 1, q.get("disciplina") or "")
        ws.cell(i, 2, q.get("item_code") or "")
        ws.cell(i, 3, q.get("descricao") or "")
        ws.cell(i, 4, q.get("unidade") or "")
        ws.cell(i, 5, _to_float(q.get("quantidade")))
        ws.cell(i, 6, q.get("calculo_memorial") or "")
        ws.cell(i, 7, q.get("origem_ambiente") or "")
        conf = _to_float(q.get("confidence"))
        c = ws.cell(i, 8, conf)
        c.number_format = "0.0%"
        ws.cell(i, 9, "SIM" if q.get("needs_review") else "")
        ws.cell(i, 5).number_format = "#,##0.00"
        for col in range(1, len(headers) + 1):
            ws.cell(i, col).border = _BORDER

    ws.freeze_panes = "B2"


def _aba_curva_abc(wb: Workbook, items: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Curva ABC")
    headers = ["EAP", "Descrição", "Custo Total (R$)", "Peso (%)", "Acumulado (%)", "Classe"]
    widths = [12, 50, 18, 12, 14, 10]
    _write_table_header(ws, headers, widths)

    sorted_items = sorted(
        (it for it in items if _to_float(it.get("custo_total")) > 0),
        key=lambda it: _to_float(it.get("custo_total")),
        reverse=True,
    )
    total = sum(_to_float(it.get("custo_total")) for it in sorted_items)

    if total <= 0:
        ws.cell(2, 1, "Sem itens com custo > 0 ainda. Rode atualizar_curva_abc(project_id) primeiro.")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        return

    acc = 0.0
    for i, it in enumerate(sorted_items, start=2):
        custo = _to_float(it.get("custo_total"))
        peso = custo / total * 100
        acc += peso
        classe = "A" if acc <= 80 else "B" if acc <= 95 else "C"

        ws.cell(i, 1, it.get("eap_code") or "")
        ws.cell(i, 2, it.get("descricao") or "")
        ws.cell(i, 3, custo)
        ws.cell(i, 4, peso / 100)
        ws.cell(i, 5, acc / 100)
        ws.cell(i, 6, classe)

        ws.cell(i, 3).number_format = "#,##0.00"
        ws.cell(i, 4).number_format = "0.00%"
        ws.cell(i, 5).number_format = "0.00%"
        for col in range(1, len(headers) + 1):
            ws.cell(i, col).border = _BORDER

        fill = (
            _CURVA_A_FILL if classe == "A"
            else _CURVA_B_FILL if classe == "B"
            else _CURVA_C_FILL
        )
        for col in range(1, len(headers) + 1):
            ws.cell(i, col).fill = fill

    ws.freeze_panes = "B2"


def _write_table_header(ws: Worksheet, headers: list[str], widths: list[int]) -> None:
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(1, col, h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = w


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0
